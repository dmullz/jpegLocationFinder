from PIL import Image
from PIL.ExifTags import TAGS
from PIL.ExifTags import GPSTAGS
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
import requests
import os
import struct

HERE_APP_ID = "vQCC4M7DiyR3w4sNySuk"
HERE_APP_CODE = "J0hDHbYEc9lf5WXlXszxCw"

# @DEV - Function to add a row to the excel file
# @param excel - string of file location of excel document
# @param filename - string of path to an image file
# @param location - string of the zip code parsed from image file
def add_to_excel(excel,filename,location):
	# Format row
	new_row = [filename,location]
        
    # Load excel or build excel if this is the first one
	try:
		workbook = load_workbook(excel)
		worksheet = workbook.worksheets[0]
	except FileNotFoundError:
		headers_row = ["Image File", "Zip Code"]
		font = Font(bold=True)
		workbook = Workbook()
		worksheet = workbook.active
		worksheet.append(headers_row)
		for cell in worksheet["1:1"]:
			cell.font = font
	
	# Add row to Excel doc
	worksheet.append(new_row)
	workbook.save(excel)

# @DEV - Generator to get the next image file
# @param dir - directory containing image files (or singular file)
# @RET (string) - Yields a file to attempt to decode
def image_gen(dir):
	# For single files, simply yield the file
	if os.path.isfile(dir):
		yield dir
	else:
		# Walk the dir structure and extract the filenames
		files = []
		for (dirpath, dirnames, filenames) in os.walk(dir):
			files.extend(filenames)
			break
		# Yield all files found
		for file in files:
			yield file

# @DEV - Function to find geotagging data in EXIF data
# @param exif - dictionary of EXIF data from a JPEG image
# @RET (dict) - returns a dictionary of geotagging data
def get_geotagging(exif):
	if not exif:
		raise ValueError("No EXIF metadata found")
	# Build geotag dictionary from number tags in EXIF data into meaningful string tags
	geotagging = {}
	for (idx, tag) in TAGS.items():
		if tag == 'GPSInfo':
			if idx not in exif:
				return None
			for (key, val) in GPSTAGS.items():
				if key in exif[idx]:
					geotagging[val] = exif[idx][key]

	return geotagging

# @DEV - Function to find geotagging data in raw EXIF data (no libs)
# @param raw - raw EXIF data
# @RET (dict) - returns a dictionary of geotagging data
def get_geotag_manual(raw):
	# check if exif data is properly formatted and extract necessary parsing vals
	if len(raw) > 15 and struct.unpack('4s',raw[:4])[0].decode('ascii') == "Exif":
		if struct.unpack('2s',raw[6:8])[0].decode('ascii') == "MM":
			byte_order = '>'
		else:
			byte_order = '<'
		header_offset = struct.unpack(byte_order + 'l',raw[10:14])[0]
	else:
		return None
	
	# Start looking through 0th IFD for GPSInfo tag
	i = 6 + header_offset
	gps_offset = 0
	while i < len(raw):
		if raw[i] == 136:
			if i+1 < len(raw) and raw[i+1] == 37:
				gps_offset = struct.unpack(byte_order+'l',raw[i+8:i+12])[0]
				break
		i += 1
	
	# Return None if no GPSInfo tag is found
	if gps_offset == 0:
		return None
	
	# Build geotag dictionary from data found in offset
	geotags = {}
	geotags["GPSVersionID"] = raw[gps_offset+header_offset+8:gps_offset+header_offset+12]
	next = 12
	while struct.unpack(byte_order+'h',raw[gps_offset + header_offset + next:gps_offset + header_offset + next+2])[0] in range(5):
		get_rational = False
		tag_num = struct.unpack(byte_order+'h',raw[gps_offset + header_offset + next:gps_offset + header_offset + next+2])[0]
		type = struct.unpack(byte_order+'h',raw[gps_offset + header_offset + next+2:gps_offset + header_offset + next+4])[0]
		length = struct.unpack(byte_order+'l',raw[gps_offset + header_offset + next+4:gps_offset + header_offset + next+8])[0]
		if type == 2:
			t='cxxx'
		if type == 3:
			t='h'
		if type == 4:
			t='l'
		if type == 5:
			t='l'
			get_rational = True
		# Extract and unpack the data field
		data = struct.unpack(byte_order+t,raw[gps_offset + header_offset + next+8:gps_offset + header_offset + next+12])[0]
		if get_rational == False:
			# For ascii values and longs, store the extracted data
			if 'c' in t:
				geotags[GPSTAGS[tag_num]] = data.decode('ascii')
			else:
				geotags[GPSTAGS[tag_num]] = data
		else:
			# Get Rational Data by unpacking data found in offset (data field)
			(d1,d2,m1,m2,s1,s2) = struct.unpack(byte_order + length*'ii', raw[data+6:data+6+(8*length)])
			geotags[GPSTAGS[tag_num]] = ((d1,d2),(m1,m2),(s1,s2))
		next += 12
	return geotags

# @DEV - Function to get the EXIF data in an image file.
# @param filename - string containing the path to a file
# @RET (dict) - returns a dictionary of EXIF data, or None if the file cannot be read or is not a JPEG
def get_exif(filename):
	# Attempt to open the file and verify it is an image
	try:
		image = Image.open(filename)
		image.verify()
		# Check if image type is jpeg
		if image.format.lower() != 'jpeg':
			print("File", filename, "is not a JPEG.")
			return None
		# Get parsed exif data
		exif = image._getexif()
		return exif
	except:
		print("There was a problem reading file", filename)
		return None

# @DEV - Function to get geotag data from a jpeg. First checks if the file is a jpeg, then attempts to extract geotag data
# @param filename - string containing the path to a file
# @RET (dict) - returns a dictionary of GEOTAG data, or None
def get_geotags_from_jpeg(filename):
	# Attempt to open the file and verify it is an image
	try:
		image = Image.open(filename)
		image.verify()
		# Check if image type is jpeg
		if image.format.lower() != 'jpeg':
			print("File", filename, "is not a JPEG.")
			return None
		# Get geotags from raw exif data
		geotags = get_geotag_manual(image.info["exif"])
		return geotags
	except:
		print("There was a problem reading file", filename)
		return None

# @DEV - Function to convert degrees, minutes, and seconds to a latitude/longitude
# @param dms - matrix of degrees, minutes, and seconds data
# @param ref - string of bearing information (N,S,E,W)
# @RET (float) - returns a float representing latitude/longitude coordinates
def get_decimal_from_dms(dms, ref):
	# degrees, minutes, seconds are expressed in rationals (numerator/denominator tuples)
	degrees = dms[0][0] / dms[0][1]
	minutes = dms[1][0] / dms[1][1] / 60.0
	seconds = dms[2][0] / dms[2][1] / 3600.0

	# flip values to negative for southern or western hemispheres
	if ref in ['S', 'W']:
		degrees = -degrees
		minutes = -minutes
		seconds = -seconds
	
	# get lat/long value by adding dms together
	return round(degrees + minutes + seconds, 5)

# @DEV - Function to get latitude/longitude from geotagging data
# @param geotags - dictionary of geotagging data
# @RET (tuple) - returns floats of latitude and longitude coordinates
def get_coordinates(geotags):
	lat = get_decimal_from_dms(geotags['GPSLatitude'], geotags['GPSLatitudeRef'])

	lon = get_decimal_from_dms(geotags['GPSLongitude'], geotags['GPSLongitudeRef'])

	return (lat,lon)

# @DEV - Function to find location data from geotagging data in a JPEG
# @param geotags - dictionary of geotagging data
# @RET (dictionary) - returns JSON containing location data for the image
def get_location(geotags):
	# Get lat/long coordinates from geotags
	coords = get_coordinates(geotags)
	
	# Use HERE API to find location data from lat/long coords
	uri = 'https://reverse.geocoder.api.here.com/6.2/reversegeocode.json'
	headers = {}
	params = {
		'app_id': HERE_APP_ID,
		'app_code': HERE_APP_CODE,
		'prox': "%s,%s" % coords,
		'gen': 9,
		'mode': 'retrieveAddresses',
		'maxresults': 1,
	}

	response = requests.get(uri, headers=headers, params=params)
	try:
		response.raise_for_status()
		return response.json()

	except requests.exceptions.HTTPError as e:
		print(str(e))
		return {}